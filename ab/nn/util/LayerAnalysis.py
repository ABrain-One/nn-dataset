"""
Layer-wise analysis for pruning and architecture search.
"""

import math
import torch
import torch.nn as nn
from typing import Dict, Optional

try:
    import weightwatcher as ww
    _HAS_WW = True
except ImportError:
    _HAS_WW = False


_WEIGHT_LAYER_TYPES = (
    nn.Conv1d, nn.Conv2d, nn.Conv3d,
    nn.ConvTranspose1d, nn.ConvTranspose2d, nn.ConvTranspose3d,
    nn.Linear,
)


def _safe_float(val, default=None):
    # Returns None for missing/NaN/inf — maps cleanly to SQL NULL
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except:
        return default


def _effective_rank_from_sv(sv: torch.Tensor) -> float:
    sv = sv[sv > 0]
    if sv.numel() == 0:
        return 0.0
    p = sv / sv.sum()
    entropy = -(p * p.log()).sum().item()
    return math.exp(entropy)


def _linear_cka(X: torch.Tensor, Y: torch.Tensor) -> float:
    X = X - X.mean(dim=0, keepdim=True)
    Y = Y - Y.mean(dim=0, keepdim=True)
    cross = torch.linalg.norm(Y.T @ X, ord='fro').item() ** 2
    norm_x = torch.linalg.norm(X.T @ X, ord='fro').item()
    norm_y = torch.linalg.norm(Y.T @ Y, ord='fro').item()
    denom = norm_x * norm_y
    return cross / denom if denom > 0 else 0.0


class LayerAnalyzer:

    _HEADERS = [
        'Layer Type', 'Layer Name', 'WW Alpha', 'Grad Norm',        # CHANGED: 'Unique ID' → 'Layer Type'
        'Dead Frac', 'Taylor Imp', 'CKA Redund', 'Eff Rank', 'Rank Ratio', 'Sensitivity',
    ]
    _COL_W = [12, 18, 10, 10, 10, 12, 12, 10, 10, 12]

    def __init__(self, model: nn.Module, device: Optional[torch.device] = None):
        self.model = model
        self.device = device or next(model.parameters()).device

    def _weight_layers(self):
        return {
            name: mod for name, mod in self.model.named_modules()
            if isinstance(mod, _WEIGHT_LAYER_TYPES)
        }

    def _get_layer_types(self):
        # CHANGED: renamed from _get_layer_ids(), removed counter logic
        # Returns just conv/linear/deconv per layer — counter was redundant with layer_idx in DB
        result = {}
        for name, mod in self._weight_layers().items():
            if isinstance(mod, nn.Linear):
                result[name] = "linear"
            elif isinstance(mod, (nn.ConvTranspose1d, nn.ConvTranspose2d, nn.ConvTranspose3d)):
                result[name] = "deconv"
            else:
                result[name] = "conv"
        return result

    def weightwatcher_analysis(self):
        if not _HAS_WW:
            return {"error": "weightwatcher not installed"}

        # Collect modules whose weight is None (e.g. CrossEntropyLoss)
        # into a list FIRST — never mutate during iteration.
        to_remove = []
        for name, mod in self.model.named_modules():
            if mod is self.model or isinstance(mod, _WEIGHT_LAYER_TYPES):
                continue
            if getattr(mod, "weight", "MISSING") is None:
                to_remove.append(name)

        # Detach them from the model so WeightWatcher doesn't see them.
        detached = []
        for name in to_remove:
            parent_name, _, attr = name.rpartition(".")
            parent = self.model.get_submodule(parent_name) if parent_name else self.model
            if hasattr(parent, attr):
                detached.append((parent, attr, getattr(parent, attr)))
                delattr(parent, attr)

        # Run WeightWatcher, then restore detached modules no matter what.
        try:
            watcher = ww.WeightWatcher(model=self.model)
            details = watcher.analyze()
        except Exception as e:
            return {"error": f"watcher.analyze() failed: {e}"}
        finally:
            for parent, attr, mod in detached:
                setattr(parent, attr, mod)

        # Match WeightWatcher rows to torch layer names by type + occurrence order.
        ww_rows = []
        for idx, row in details.iterrows():
            layer_type = str(row.get("layer_type", "")).lower()
            t = "conv" if "conv" in layer_type else "linear"
            ww_rows.append({
                "idx": int(idx),
                "type": t,
                "alpha": _safe_float(row.get("alpha")),
                "stable_rank": _safe_float(row.get("stable_rank")),
            })

        ww_rows.sort(key=lambda x: x["idx"])

        torch_layers = []
        type_counter = {}
        for name, mod in self._weight_layers().items():
            t = "linear" if isinstance(mod, nn.Linear) else "conv"
            type_counter[t] = type_counter.get(t, 0) + 1
            torch_layers.append({"name": name, "type": t, "occ": type_counter[t]})

        ww_counter = {}
        for w in ww_rows:
            t = w["type"]
            ww_counter[t] = ww_counter.get(t, 0) + 1
            w["occ"] = ww_counter[t]

        result = {l["name"]: {"alpha": None, "stable_rank": None} for l in torch_layers}

        used = set()
        for w in ww_rows:
            for i, l in enumerate(torch_layers):
                if i in used:
                    continue
                if l["type"] == w["type"] and l["occ"] == w["occ"]:
                    result[l["name"]] = {"alpha": w["alpha"], "stable_rank": w["stable_rank"]}
                    used.add(i)
                    break

        return result

    def gradient_flow(self, loader, loss_fn, num_batches=1):
        was_training = self.model.training
        self.model.train()
        grads = {}
        try:
            for i, (x, y) in enumerate(loader):
                if i >= num_batches:
                    break
                x, y = x.to(self.device), y.to(self.device)
                self.model.zero_grad()
                loss = loss_fn(self.model(x), y)
                loss.backward()
                for name, p in self.model.named_parameters():
                    if p.grad is not None:
                        grads[name] = {"grad_norm": p.grad.norm().item()}
        finally:
            self.model.zero_grad()
            self.model.train(was_training)
        return grads

    def activation_statistics(self, loader, num_batches=1):
        was_training = self.model.training
        self.model.eval()
        stats = {}
        hooks = []

        def hook(name):
            def fn(m, i, o):
                if isinstance(o, torch.Tensor):
                    flat = o.view(o.shape[0], -1)
                    stats[name] = {
                        "dead_neuron_fraction": (flat.abs() < 1e-6).float().mean().item()
                    }
            return fn

        for name, mod in self._weight_layers().items():
            hooks.append(mod.register_forward_hook(hook(name)))

        try:
            with torch.no_grad():
                for i, (x, _) in enumerate(loader):
                    if i >= num_batches:
                        break
                    self.model(x.to(self.device))
        finally:
            for h in hooks:
                h.remove()
            self.model.train(was_training)

        return stats

    def taylor_importance(self, data_loader, loss_fn, num_batches=1):
        was_training = self.model.training
        self.model.train()

        activations = {}
        hooks = []

        def _hook(name):
            def fn(module, inp, out):
                if isinstance(out, torch.Tensor):
                    cloned = out.clone()
                    cloned.retain_grad()
                    activations[name] = cloned
                    return cloned
            return fn

        for name, mod in self._weight_layers().items():
            hooks.append(mod.register_forward_hook(_hook(name)))

        importance = {}
        count = 0

        try:
            for inputs, labels in data_loader:
                if count >= num_batches:
                    break
                inputs, labels = inputs.to(self.device), labels.to(self.device)
                self.model.zero_grad()
                loss = loss_fn(self.model(inputs), labels)
                loss.backward()

                for name, act in activations.items():
                    if act.grad is None:
                        continue
                    grad = act.grad
                    if act.dim() == 4:
                        score = (act * grad).abs().mean(dim=(0, 2, 3))
                    else:
                        score = (act * grad).abs().mean(dim=0)
                    if name in importance:
                        importance[name] += score.detach()
                    else:
                        importance[name] = score.detach().clone()

                activations.clear()
                count += 1
        finally:
            for h in hooks:
                h.remove()
            self.model.zero_grad()
            self.model.train(was_training)

        result = {}
        for name, scores in importance.items():
            scores = scores / max(count, 1)
            result[name] = {
                "mean_importance": scores.mean().item(),
                "max_importance": scores.max().item(),
                "min_importance": scores.min().item(),
            }

        return result

    def _cka_per_layer(self, cka_result):
        names = cka_result["layer_names"]
        matrix = cka_result["cka_matrix"]
        n = len(names)
        per_layer = {}
        for i, name in enumerate(names):
            vals = [matrix[i][j] for j in range(n) if j != i]
            per_layer[name] = sum(vals) / len(vals) if vals else None
        return per_layer

    def cka_similarity(self, loader, num_batches=1):
        was_training = self.model.training
        self.model.eval()
        reps = {}
        hooks = []

        def hook(name):
            def fn(m, i, o):
                if isinstance(o, torch.Tensor):
                    a = o.detach()
                    if a.dim() == 4:
                        a = a.mean((2, 3))
                    reps[name] = a.cpu()
            return fn

        for name, mod in self._weight_layers().items():
            hooks.append(mod.register_forward_hook(hook(name)))

        try:
            with torch.no_grad():
                for i, (x, _) in enumerate(loader):
                    if i >= num_batches:
                        break
                    self.model(x.to(self.device))
        finally:
            for h in hooks:
                h.remove()
            self.model.train(was_training)

        names = list(reps)
        n = len(names)
        mat = [[_linear_cka(reps[names[i]], reps[names[j]]) for j in range(n)] for i in range(n)]

        return {"layer_names": names, "cka_matrix": mat}

    def effective_rank(self):
        result = {}
        for name, mod in self._weight_layers().items():
            w = mod.weight.data.reshape(mod.weight.shape[0], -1)
            sv = torch.linalg.svdvals(w)
            eff = _effective_rank_from_sv(sv)
            result[name] = {
                "effective_rank": eff,
                "rank_ratio": eff / min(w.shape),
            }
        return result

    def sensitivity_analysis(self, data_loader, loss_fn, num_batches=1):
        was_training = self.model.training
        self.model.eval()

        baseline = 0.0
        batches = []
        count = 0

        with torch.no_grad():
            for inputs, labels in data_loader:
                if count >= num_batches:
                    break
                batches.append((inputs.cpu(), labels.cpu()))
                inputs, labels = inputs.to(self.device), labels.to(self.device)
                baseline += loss_fn(self.model(inputs), labels).item()
                count += 1

        baseline /= max(count, 1)

        result = {}
        for name, mod in self._weight_layers().items():
            orig_w = mod.weight.data.clone()
            orig_b = mod.bias.data.clone() if mod.bias is not None else None
            try:
                mod.weight.data.zero_()
                if mod.bias is not None:
                    mod.bias.data.zero_()
                zeroed = 0.0
                with torch.no_grad():
                    for inp, lab in batches:
                        zeroed += loss_fn(self.model(inp.to(self.device)), lab.to(self.device)).item()
                zeroed /= max(count, 1)
                result[name] = {
                    "relative_sensitivity": (zeroed - baseline) / max(abs(baseline), 1e-8)
                }
            except Exception:
                pass
            finally:
                mod.weight.data.copy_(orig_w)
                if orig_b is not None:
                    mod.bias.data.copy_(orig_b)

        self.model.train(was_training)
        return result

    def full_analysis(self, loader, loss_fn, num_batches=1):
        methods = [
            ("weightwatcher",         lambda: self.weightwatcher_analysis()),
            ("gradient_flow",         lambda: self.gradient_flow(loader, loss_fn, num_batches)),
            ("activation_statistics", lambda: self.activation_statistics(loader, num_batches)),
            ("taylor_importance",     lambda: self.taylor_importance(loader, loss_fn, num_batches)),
            ("effective_rank",        lambda: self.effective_rank()),
            ("cka_similarity",        lambda: self.cka_similarity(loader, num_batches)),
            ("sensitivity",           lambda: self.sensitivity_analysis(loader, loss_fn, num_batches=5)),
        ]
        result = {}
        for name, fn in methods:
            try:
                result[name] = fn()
            except Exception as e:
                result[name] = {"error": str(e)}
        result["layer_types"] = self._get_layer_types()
        # CHANGED: was result["layer_ids"] = self._get_layer_ids()
        return result

    def build_layer_table(self, analysis):
        types   = analysis.get("layer_types", {})       # CHANGED: was layer_ids
        ww      = analysis.get("weightwatcher", {})
        gf      = analysis.get("gradient_flow", {})
        act     = analysis.get("activation_statistics", {})
        ti      = analysis.get("taylor_importance", {})
        er      = analysis.get("effective_rank", {})
        sa      = analysis.get("sensitivity", {})
        cka_raw = analysis.get("cka_similarity", {})
        cka     = self._cka_per_layer(cka_raw) if "cka_matrix" in cka_raw else {}

        # CHANGED: removed _na() helper entirely
        # All missing/NaN/inf values are now None, which maps to SQL NULL
        # _safe_float() handles NaN/inf → None for numeric fields
        # String fields use plain dict.get() which returns None if missing

        table = {}
        for name in self._weight_layers().keys():
            table[name] = {
                "layer_type": types.get(name),                                          # CHANGED: was unique_id with _na()
                "ww_alpha":   _safe_float(ww.get(name, {}).get("alpha")),               # CHANGED: was _na()
                "grad_norm":  _safe_float(gf.get(name + ".weight", {}).get("grad_norm")),
                "dead_frac":  _safe_float(act.get(name, {}).get("dead_neuron_fraction")),
                "taylor_imp": _safe_float(ti.get(name, {}).get("mean_importance")),
                "cka_redund": _safe_float(cka.get(name)),
                "eff_rank":   _safe_float(er.get(name, {}).get("effective_rank")),
                "rank_ratio": _safe_float(er.get(name, {}).get("rank_ratio")),
                "sensitivity":_safe_float(sa.get(name, {}).get("relative_sensitivity")),
            }
        return table

    @staticmethod
    def summarize(analysis):
        summary = {}

        ww = analysis.get("weightwatcher", {})
        alphas = [v.get("alpha") for v in ww.values()
                  if isinstance(v, dict) and isinstance(v.get("alpha"), (int, float))]
        if alphas:
            summary["ww_mean_alpha"] = sum(alphas) / len(alphas)

        gf = analysis.get("gradient_flow", {})
        norms = [v["grad_norm"] for k, v in gf.items()
                 if k.endswith(".weight") and isinstance(v, dict) and "grad_norm" in v]
        if norms:
            summary["gf_mean_grad_norm"] = sum(norms) / len(norms)

        act = analysis.get("activation_statistics", {})
        deads = [v["dead_neuron_fraction"] for v in act.values()
                 if isinstance(v, dict) and "dead_neuron_fraction" in v]
        if deads:
            summary["as_dead_neuron_frac"] = sum(deads) / len(deads)

        ti = analysis.get("taylor_importance", {})
        imps = [v["mean_importance"] for v in ti.values()
                if isinstance(v, dict) and "mean_importance" in v]
        if imps:
            summary["ti_mean_importance"] = sum(imps) / len(imps)

        cka = analysis.get("cka_similarity", {})
        if "cka_matrix" in cka:
            matrix, n = cka["cka_matrix"], len(cka["cka_matrix"])
            if n > 1:
                off = [matrix[i][j] for i in range(n) for j in range(i + 1, n)]
                if off:
                    summary["cka_mean_redundancy"] = sum(off) / len(off)

        er = analysis.get("effective_rank", {})
        ratios = [v["rank_ratio"] for v in er.values()
                  if isinstance(v, dict) and "rank_ratio" in v]
        if ratios:
            summary["er_mean_rank_ratio"] = sum(ratios) / len(ratios)

        sa = analysis.get("sensitivity", {})
        sens = [v["relative_sensitivity"] for v in sa.values()
                if isinstance(v, dict) and "relative_sensitivity" in v]
        if sens:
            summary["sa_mean_sensitivity"] = sum(sens) / len(sens)

        return summary

    @staticmethod
    def _fmt(v, decimals=4):
        if v is None or v == 'N/A':
            return 'N/A'
        if not isinstance(v, (int, float)):
            return str(v)
        if abs(v) < 0.001 and v != 0.0:
            return f'{v:.2e}'
        return f'{v:.{decimals}f}'

    @staticmethod
    def view(table):
        h, w = LayerAnalyzer._HEADERS, LayerAnalyzer._COL_W
        print(' | '.join(hh.ljust(ww) for hh, ww in zip(h, w)))
        print('-+-'.join('-' * ww for ww in w))
        for layer_name, row in table.items():
            vals = [
                row.get('layer_type',  'N/A'),      # CHANGED: was unique_id
                layer_name,
                row.get('ww_alpha',   'N/A'),
                row.get('grad_norm',  'N/A'),
                row.get('dead_frac',  'N/A'),
                row.get('taylor_imp', 'N/A'),
                row.get('cka_redund', 'N/A'),
                row.get('eff_rank',   'N/A'),
                row.get('rank_ratio', 'N/A'),
                row.get('sensitivity','N/A'),
            ]
            print(' | '.join(LayerAnalyzer._fmt(v).ljust(ww) for v, ww in zip(vals, w)))

    @staticmethod
    def export(table, out_path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("[ERROR] export() requires openpyxl: pip install openpyxl")
            return

        h = LayerAnalyzer._HEADERS
        col_widths = [12, 18, 11, 11, 11, 12, 12, 11, 11, 13]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Layer Analysis'

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color='2F4F8F')
        alt_fill    = PatternFill('solid', start_color='EEF2FF')

        ws.row_dimensions[1].height = 28
        for col, hh in enumerate(h, 1):
            cell = ws.cell(row=1, column=col, value=hh)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[cell.column_letter].width = col_widths[col - 1]

        for row_idx, (layer_name, row) in enumerate(table.items(), 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            vals = [
                row.get('layer_type',  'N/A'),      # CHANGED: was unique_id
                layer_name,
                row.get('ww_alpha',   'N/A'),
                row.get('grad_norm',  'N/A'),
                row.get('dead_frac',  'N/A'),
                row.get('taylor_imp', 'N/A'),
                row.get('cka_redund', 'N/A'),
                row.get('eff_rank',   'N/A'),
                row.get('rank_ratio', 'N/A'),
                row.get('sensitivity','N/A'),
            ]
            for col_idx, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Arial', size=10,
                                 color='999999' if v == 'N/A' else '000000',
                                 italic=(v == 'N/A'))
                if fill:
                    cell.fill = fill

        try:
            wb.save(out_path)
            print(f"Exported: {out_path}")
        except Exception as e:
            print(f"[ERROR] Failed to export xlsx: {e}")


if __name__ == '__main__':
    import json
    import sys
    from pathlib import Path

    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path or not Path(path).exists():
        print("Usage: python -m ab.nn.util.LayerAnalysis <path_to_layer_table.json> [--export]")
        sys.exit(1)

    with open(path) as f:
        table = json.load(f)

    print(f"\nLayer Table  —  {Path(path).name}")
    print('=' * 120)
    LayerAnalyzer.view(table)

    if '--export' in sys.argv:
        out = Path(path).parent / (Path(path).stem + '_export.xlsx')
        LayerAnalyzer.export(table, str(out))